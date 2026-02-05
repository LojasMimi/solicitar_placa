import streamlit as st
import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
from io import BytesIO
from datetime import datetime
import time

# ===============================
# CONFIGURAÇÃO DA PÁGINA
# ===============================
st.set_page_config(
    page_title="Solicitação de placas",
    page_icon="🏷️",
    layout="centered"
)

st.title("Solicitação de placas")

# ===============================
# CONSTANTES
# ===============================
LIMITE_PRODUTOS = 23  # Limite máximo de produtos permitidos

# ===============================
# SESSION STATE
# ===============================
if "workbook" not in st.session_state:
    st.session_state.workbook = None

if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None

if "produtos" not in st.session_state:
    st.session_state.produtos = []

if "loja" not in st.session_state:
    st.session_state.loja = ""

if "nome_solicitante" not in st.session_state:
    st.session_state.nome_solicitante = ""

if "data_solicitacao" not in st.session_state:
    st.session_state.data_solicitacao = ""

# Variáveis para controlar exibição de mensagens
if "mensagem_sucesso" not in st.session_state:
    st.session_state.mensagem_sucesso = None
if "mensagem_lote" not in st.session_state:
    st.session_state.mensagem_lote = None
if "mensagem_produto" not in st.session_state:
    st.session_state.mensagem_produto = None

# ===============================
# MAPEAMENTOS
# ===============================
TIPO_PLACA_MAP = {
    "1 - OFERTA": 1,
    "2 - PROMOÇÃO": 2,
    "3 - SUPER OFERTA": 3,
    "4 - MEGA OFERTA": 4,
    "5 - SINALIZAÇÃO SIMPLES": 5
}

TAMANHO_PLACA_MAP = {
    "A - FOLHA HORIZONTAL": "A",
    "B - FOLHA VERTICAL": "B",
    "C - MEIA FOLHA": "C",
    "D - 1/4 FOLHA": "D",
    "E - PORTA ETIQUETA": "E"
}

TAMANHO_PLACA_IMAGEM_MAP = {
    "A - FOLHA HORIZONTAL": "imagens/HORIZONTAL.png",
    "B - FOLHA VERTICAL": "imagens/VERTICAL.png",
    "C - MEIA FOLHA": "imagens/MEIA_FOLHA.jpg",
    "D - 1/4 FOLHA": "imagens/UM_QUARTO_FOLHA.jpg",
    "E - PORTA ETIQUETA": "imagens/ETIQUETA_GONDOLA.JFIF"
}

# ===============================
# FUNÇÕES AUXILIARES
# ===============================
def criar_planilha_from_scratch():
    """Cria uma nova planilha com todos os produtos do session state"""
    wb = load_workbook("solicitar placa.xlsx")
    ws = wb.active
    
    # Limpar todas as linhas de produtos (da linha 16 em diante)
    for row in range(16, ws.max_row + 1):
        for col in range(2, 9):  # Colunas B a H
            ws.cell(row=row, column=col).value = None
    
    # Preencher cabeçalho
    ws["C7"].value = st.session_state.loja
    ws["D7"].value = f"DATA: {st.session_state.data_solicitacao}"
    ws["B9"].value = f"SOLICITANTE: {st.session_state.nome_solicitante}"
    
    # Preencher produtos (máximo LIMITE_PRODUTOS)
    for i, produto in enumerate(st.session_state.produtos[:LIMITE_PRODUTOS]):
        linha = 16 + i
        ws.cell(row=linha, column=2).value = produto["Tipo Placa"]  # Coluna B
        ws.cell(row=linha, column=3).value = produto["Tamanho Placa"]  # Coluna C
        ws.cell(row=linha, column=4).value = produto["Fornecedor"]   # Coluna D
        ws.cell(row=linha, column=5).value = produto["Código de Barras"]  # Coluna E
        ws.cell(row=linha, column=6).value = produto["Identificador de Origem"]  # Coluna F
        ws.cell(row=linha, column=7).value = produto["Descrição"]  # Coluna G
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    st.session_state.workbook = wb
    st.session_state.excel_buffer = buffer
    
    return buffer

def verificar_limite_produtos():
    """Verifica se atingiu o limite de produtos e retorna mensagem se necessário"""
    quantidade_atual = len(st.session_state.produtos)
    
    if quantidade_atual >= LIMITE_PRODUTOS:
        produtos_excedentes = quantidade_atual - LIMITE_PRODUTOS
        return f"⚠️ LIMITE ATINGIDO! O formulário suporta apenas {LIMITE_PRODUTOS} produtos. Você já tem {quantidade_atual} produtos ({produtos_excedentes} excedentes)."
    
    return None

# ===============================
# ABAS
# ===============================
tab_individual, tab_lote, tab_relatorio = st.tabs(["📝 INDIVIDUAL", "📦 LOTE", "📊 RELATÓRIO"])

# ======================================================
# ABA INDIVIDUAL (antiga SOLICITAR)
# ======================================================
with tab_individual:
    st.subheader("Dados do solicitante")

    col1, col2 = st.columns([2, 1])

    with col1:
        nome_solicitante = st.text_input("Nome do solicitante")

    with col2:
        loja = st.selectbox("Loja", ["MIMI", "KAMI", "TOTAL MIX"])

    # Exibir contador de produtos atual
    quantidade_atual = len(st.session_state.produtos)
    st.info(f"📊 **Produtos solicitados:** {quantidade_atual}/{LIMITE_PRODUTOS}")
    
    if quantidade_atual >= LIMITE_PRODUTOS:
        st.warning(f"🚫 **LIMITE ATINGIDO!** Você já atingiu o limite máximo de {LIMITE_PRODUTOS} produtos.")

    # Botão para iniciar solicitação
    if st.button("Iniciar solicitação", key="iniciar_individual"):
        if not nome_solicitante.strip():
            st.warning("⚠️ Informe o nome do solicitante.")
            st.session_state.mensagem_sucesso = None
        else:
            wb = load_workbook("solicitar placa.xlsx")
            ws = wb.active

            ws["C7"].value = loja
            ws["D7"].value = f"DATA: {datetime.now().strftime('%d/%m/%Y')}"
            ws["E7"].value = f"SOLICITANTE: {nome_solicitante}"
            
            # Salvar informações no session state
            st.session_state.loja = loja
            st.session_state.nome_solicitante = nome_solicitante
            st.session_state.data_solicitacao = datetime.now().strftime('%d/%m/%Y')

            # Criar buffer da planilha
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            st.session_state.workbook = wb
            st.session_state.excel_buffer = buffer
            st.session_state.produtos = []
            st.session_state.mensagem_sucesso = "✅ Solicitação iniciada com sucesso!"

    # Exibir mensagem de sucesso se existir
    if st.session_state.mensagem_sucesso:
        st.success(st.session_state.mensagem_sucesso)
        # Limpar após exibir
        st.session_state.mensagem_sucesso = None

    st.markdown("---")
    st.subheader("Consulta de produto")

    col1, col2, col3 = st.columns([2, 2, 2])
    
    with col1:
        codigo_barras = st.text_input("Digite o código de barras", key="codigo_individual")
    
    with col2:
        tipo_placa = st.selectbox("Tipo de placa", list(TIPO_PLACA_MAP.keys()), key="tipo_individual")
    
    with col3:
        tamanho_placa = st.selectbox("Tamanho da placa", list(TAMANHO_PLACA_MAP.keys()), key="tamanho_individual")
    
    # Mostrar imagem do tamanho selecionado
    if tamanho_placa in TAMANHO_PLACA_IMAGEM_MAP:
        imagem_path = TAMANHO_PLACA_IMAGEM_MAP[tamanho_placa]
        try:
            col_img1, col_img2, col_img3 = st.columns([1, 2, 1])
            with col_img2:
                st.markdown(f"**Visualização:** {tamanho_placa.split(' - ')[1]}")
                st.image(imagem_path, width=300)
        except Exception as e:
            st.warning(f"Imagem não encontrada: {imagem_path}")
            st.info("Verifique se o arquivo existe na pasta 'imagens/'")

    # Botão para consultar produto
    if st.button("Consultar Produto", key="consultar_individual"):
        if not st.session_state.workbook:
            st.error("❌ Inicie a solicitação antes.")
            st.session_state.mensagem_produto = None
        elif not codigo_barras.strip():
            st.warning("⚠️ Por favor, digite um código de barras válido.")
            st.session_state.mensagem_produto = None
        elif any(p["Código de Barras"] == str(codigo_barras) for p in st.session_state.produtos):
            st.error("❌ PRODUTO JÁ SOLICITADO. POR FAVOR, COLOQUE OUTRO CÓDIGO DE BARRAS.")
            st.session_state.mensagem_produto = None
        elif len(st.session_state.produtos) >= LIMITE_PRODUTOS:
            st.error(f"🚫 LIMITE ATINGIDO! O formulário suporta apenas {LIMITE_PRODUTOS} produtos. Não é possível adicionar mais produtos.")
            st.session_state.mensagem_produto = None
        else:
            headers = {
                'x-api-key': st.secrets["api"]["x_api_key"],
                'Cookie': st.secrets["api"]["cookie"]
            }

            try:
                url_1 = f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/consulta/0{codigo_barras}"
                response_1 = requests.get(url_1, headers=headers)

                if response_1.status_code != 200:
                    st.error("❌ Produto não encontrado.")
                    st.session_state.mensagem_produto = None
                else:
                    dados_produto = response_1.json()
                    produto_id = dados_produto.get("id")
                    descricao = dados_produto.get("descricao")
                    ref = dados_produto.get("identificadorDeOrigem")

                    # Exibir mensagem de sucesso imediatamente
                    st.session_state.mensagem_produto = "✅ Produto encontrado com sucesso!"
                    
                    url_2 = f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/{produto_id}/precos"
                    requests.get(url_2, headers=headers)

                    fornecedor_nome = "Não encontrado"
                    url_3 = f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/{produto_id}/fornecedores"
                    response_3 = requests.get(url_3, headers=headers)

                    if response_3.status_code == 200:
                        items = response_3.json().get("items", [])
                        if items:
                            fornecedor_id = items[0].get("fornecedorId")
                            url_4 = f"https://lojasmimi.varejofacil.com/api/v1/pessoa/fornecedores?q=id=={fornecedor_id}"
                            response_4 = requests.get(url_4, headers=headers)
                            if response_4.status_code == 200:
                                forn_items = response_4.json().get("items", [])
                                if forn_items:
                                    fornecedor_nome = forn_items[0].get("fantasia")

                    produto_info = {
                        "Código de Barras": str(codigo_barras),
                        "Descrição": descricao,
                        "Fornecedor": fornecedor_nome,
                        "Identificador de Origem": ref,
                        "Tipo Placa": TIPO_PLACA_MAP[tipo_placa],
                        "Tamanho Placa": TAMANHO_PLACA_MAP[tamanho_placa]
                    }

                    st.session_state.produtos.append(produto_info)
                    
                    # Verificar limite após adicionar
                    limite_msg = verificar_limite_produtos()
                    if limite_msg:
                        st.warning(limite_msg)
                    
                    # Recriar a planilha do zero
                    criar_planilha_from_scratch()
                    
                    # Atualizar mensagem
                    if len(st.session_state.produtos) <= LIMITE_PRODUTOS:
                        st.session_state.mensagem_produto = "✅ Produto registrado corretamente!"
                    else:
                        st.session_state.mensagem_produto = f"⚠️ Produto registrado, mas ATENÇÃO: limite de {LIMITE_PRODUTOS} produtos excedido!"
                    
                    # Forçar rerun para atualizar a lista
                    st.rerun()

            except Exception as e:
                st.exception(e)
                st.session_state.mensagem_produto = None

    # Exibir mensagem do produto se existir
    if st.session_state.mensagem_produto:
        st.success(st.session_state.mensagem_produto)
        # Não limpamos imediatamente, deixamos visível

# ======================================================
# ABA LOTE
# ======================================================
with tab_lote:
    st.subheader("Solicitação em Lote")

    st.markdown("📥 Baixe o modelo, preencha apenas os códigos de barras e faça upload.")

    if st.button("Baixar modelo Excel", key="baixar_modelo"):
        modelo = Workbook()
        ws_modelo = modelo.active
        ws_modelo.title = "Produtos"
        ws_modelo["A1"].value = "CODIGO DE BARRAS"

        buffer_modelo = BytesIO()
        modelo.save(buffer_modelo)
        buffer_modelo.seek(0)

        st.download_button(
            "📥 Baixar modelo",
            data=buffer_modelo,
            file_name="modelo_lote.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_modelo"
        )

    st.markdown("---")

    arquivo_lote = st.file_uploader("Faça upload do arquivo Excel preenchido", type=["xlsx"], key="upload_lote")

    col_lote1, col_lote2 = st.columns(2)
    
    with col_lote1:
        tipo_placa_lote = st.selectbox("Tipo de placa (LOTE)", list(TIPO_PLACA_MAP.keys()), key="tipo_lote")
    
    with col_lote2:
        tamanho_placa_lote = st.selectbox("Tamanho da placa (LOTE)", list(TAMANHO_PLACA_MAP.keys()), key="tamanho_lote")
    
    # Mostrar imagem do tamanho selecionado na aba LOTE
    if tamanho_placa_lote in TAMANHO_PLACA_IMAGEM_MAP:
        imagem_path_lote = TAMANHO_PLACA_IMAGEM_MAP[tamanho_placa_lote]
        try:
            col_img_lote1, col_img_lote2, col_img_lote3 = st.columns([1, 2, 1])
            with col_img_lote2:
                st.markdown(f"**Visualização:** {tamanho_placa_lote.split(' - ')[1]}")
                st.image(imagem_path_lote, width=300)
        except Exception as e:
            st.warning(f"Imagem não encontrada: {imagem_path_lote}")
            st.info("Verifique se o arquivo existe na pasta 'imagens/'")

    # Exibir contador atual
    quantidade_atual = len(st.session_state.produtos)
    espaco_disponivel = max(0, LIMITE_PRODUTOS - quantidade_atual)
    st.info(f"📊 **Produtos solicitados:** {quantidade_atual}/{LIMITE_PRODUTOS} | **Espaço disponível:** {espaco_disponivel}")
    
    if quantidade_atual >= LIMITE_PRODUTOS:
        st.warning(f"🚫 **LIMITE ATINGIDO!** Você já atingiu o limite máximo de {LIMITE_PRODUTOS} produtos.")
    elif espaco_disponivel < 10:
        st.warning(f"⚠️ **ATENÇÃO:** Apenas {espaco_disponivel} espaços disponíveis no formulário.")

    # Botão para processar lote
    if st.button("Processar lote", key="processar_lote"):
        if not st.session_state.workbook:
            st.error("❌ Inicie a solicitação na aba INDIVIDUAL antes de processar o lote.")
            st.session_state.mensagem_lote = None
        elif not arquivo_lote:
            st.warning("⚠️ Faça upload de um arquivo válido.")
            st.session_state.mensagem_lote = None
        elif quantidade_atual >= LIMITE_PRODUTOS:
            st.error(f"🚫 LIMITE ATINGIDO! O formulário já está com {LIMITE_PRODUTOS} produtos. Não é possível processar mais produtos.")
            st.session_state.mensagem_lote = None
        else:
            try:
                df_lote = pd.read_excel(arquivo_lote)
                codigos = df_lote["CODIGO DE BARRAS"].astype(str).tolist()
                
                # Verificar quantos produtos podem ser adicionados
                espaco_disponivel = LIMITE_PRODUTOS - quantidade_atual
                codigos_para_processar = codigos[:espaco_disponivel]
                codigos_excedentes = codigos[espaco_disponivel:] if len(codigos) > espaco_disponivel else []
                
                # Aviso sobre produtos excedentes
                if codigos_excedentes:
                    st.warning(f"⚠️ **ATENÇÃO:** O arquivo contém {len(codigos)} produtos, mas apenas {espaco_disponivel} serão processados devido ao limite do formulário.")
                    st.info(f"📋 {len(codigos_excedentes)} produtos serão ignorados por falta de espaço.")

                produtos_sucesso = []
                produtos_falha = []
                produtos_duplicados = []

                # Criar barra de progresso
                progress_bar = st.progress(0)
                status_text = st.empty()
                total_processar = len(codigos_para_processar)

                headers = {
                    'x-api-key': st.secrets["api"]["x_api_key"],
                    'Cookie': st.secrets["api"]["cookie"]
                }

                for i, codigo in enumerate(codigos_para_processar, start=1):
                    time.sleep(0.05)
                    progresso = i / total_processar
                    progress_bar.progress(progresso)
                    status_text.text(f"Processando {i} de {total_processar} produtos...")

                    # Verificar duplicado
                    if any(p["Código de Barras"] == codigo for p in st.session_state.produtos):
                        produtos_duplicados.append({"Código de Barras": codigo, "Erro": "Duplicado"})
                        continue

                    url_1 = f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/consulta/0{codigo}"
                    response_1 = requests.get(url_1, headers=headers)

                    if response_1.status_code != 200:
                        produtos_falha.append({"Código de Barras": codigo, "Erro": "Não encontrado"})
                        continue

                    dados_produto = response_1.json()
                    produto_id = dados_produto.get("id")
                    descricao = dados_produto.get("descricao")
                    ref = dados_produto.get("identificadorDeOrigem")

                    fornecedor_nome = "Não encontrado"
                    url_3 = f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/{produto_id}/fornecedores"
                    response_3 = requests.get(url_3, headers=headers)

                    if response_3.status_code == 200:
                        items = response_3.json().get("items", [])
                        if items:
                            fornecedor_id = items[0].get("fornecedorId")
                            url_4 = f"https://lojasmimi.varejofacil.com/api/v1/pessoa/fornecedores?q=id=={fornecedor_id}"
                            response_4 = requests.get(url_4, headers=headers)
                            if response_4.status_code == 200:
                                forn_items = response_4.json().get("items", [])
                                if forn_items:
                                    fornecedor_nome = forn_items[0].get("fantasia")

                    produto_info = {
                        "Código de Barras": codigo,
                        "Descrição": descricao,
                        "Fornecedor": fornecedor_nome,
                        "Identificador de Origem": ref,
                        "Tipo Placa": TIPO_PLACA_MAP[tipo_placa_lote],
                        "Tamanho Placa": TAMANHO_PLACA_MAP[tamanho_placa_lote]
                    }

                    st.session_state.produtos.append(produto_info)
                    produtos_sucesso.append(produto_info)

                # Limpar barra de progresso
                progress_bar.empty()
                status_text.empty()

                # Recriar a planilha do zero após processar todos os produtos
                criar_planilha_from_scratch()

                # Criar mensagem de sucesso detalhada
                mensagem_detalhada = f"✅ Lote processado com sucesso!\n\n"
                mensagem_detalhada += f"📊 **Resumo:**\n"
                mensagem_detalhada += f"- ✅ Produtos adicionados: {len(produtos_sucesso)}\n"
                mensagem_detalhada += f"- ❌ Produtos não encontrados: {len(produtos_falha)}\n"
                mensagem_detalhada += f"- ⚠️ Produtos duplicados: {len(produtos_duplicados)}\n"
                
                if codigos_excedentes:
                    mensagem_detalhada += f"- 🚫 Produtos excedentes (não processados): {len(codigos_excedentes)}\n"
                
                mensagem_detalhada += f"\n📈 **Total no formulário:** {len(st.session_state.produtos)}/{LIMITE_PRODUTOS}"
                
                st.session_state.mensagem_lote = mensagem_detalhada
                
                # Exibir mensagem imediatamente
                st.success("✅ Lote processado com sucesso!")
                st.markdown(mensagem_detalhada)

                # Mostrar produtos que falharam
                if produtos_falha:
                    st.markdown("### ❌ Produtos não encontrados")
                    df_falha = pd.DataFrame(produtos_falha)
                    st.dataframe(df_falha, use_container_width=True)
                
                # Mostrar produtos duplicados
                if produtos_duplicados:
                    st.markdown("### ⚠️ Produtos duplicados (já existentes)")
                    df_duplicados = pd.DataFrame(produtos_duplicados)
                    st.dataframe(df_duplicados, use_container_width=True)
                
                # Mostrar produtos excedentes
                if codigos_excedentes:
                    st.markdown("### 🚫 Produtos excedentes (não processados por limite)")
                    df_excedentes = pd.DataFrame({
                        "Código de Barras": codigos_excedentes,
                        "Motivo": f"Limite de {LIMITE_PRODUTOS} produtos atingido"
                    })
                    st.dataframe(df_excedentes, use_container_width=True)
                    st.info(f"⚠️ **Recomendação:** Processe os {len(codigos_excedentes)} produtos restantes em um novo lote após limpar o formulário.")
                    
                # Forçar rerun para atualizar
                st.rerun()

            except Exception as e:
                st.exception(e)
                st.session_state.mensagem_lote = None

    # Exibir mensagem do lote se existir (para quando a página recarrega)
    if st.session_state.mensagem_lote:
        st.success("✅ Lote processado anteriormente")
        st.markdown(st.session_state.mensagem_lote)

# ======================================================
# ABA RELATÓRIO
# ======================================================
with tab_relatorio:
    st.subheader("Produtos solicitados")
    
    # Mostrar contador
    quantidade_atual = len(st.session_state.produtos)
    st.info(f"📊 **Total de produtos:** {quantidade_atual}/{LIMITE_PRODUTOS}")
    
    if quantidade_atual > LIMITE_PRODUTOS:
        st.error(f"🚫 **ATENÇÃO:** O formulário tem {quantidade_atual} produtos, mas suporta apenas {LIMITE_PRODUTOS}. "
                f"Os primeiros {LIMITE_PRODUTOS} serão incluídos na planilha.")

    if st.session_state.produtos:
        # Mostrar apenas os produtos que cabem no formulário
        produtos_para_exibir = st.session_state.produtos[:LIMITE_PRODUTOS]
        produtos_excedentes = st.session_state.produtos[LIMITE_PRODUTOS:] if quantidade_atual > LIMITE_PRODUTOS else []
        
        df = pd.DataFrame(produtos_para_exibir)
        st.dataframe(df, use_container_width=True)
        
        # Mostrar alerta sobre produtos excedentes
        if produtos_excedentes:
            st.warning(f"⚠️ **ATENÇÃO:** {len(produtos_excedentes)} produtos excedem o limite do formulário e NÃO serão incluídos na planilha:")
            df_excedentes = pd.DataFrame(produtos_excedentes)
            st.dataframe(df_excedentes, use_container_width=True)
            st.error(f"🚫 **Ação necessária:** Remova {len(produtos_excedentes)} produtos para não perder informações.")

        st.markdown("---")
        st.subheader("🗑️ Remover produto")

        # Criar opções para remoção (incluindo produtos excedentes)
        options = []
        for i, p in enumerate(st.session_state.produtos):
            descricao_curta = p["Descrição"][:50] + "..." if len(p["Descrição"]) > 50 else p["Descrição"]
            marcador = "🚫 " if i >= LIMITE_PRODUTOS else ""
            options.append(f'{marcador}{p["Código de Barras"]} - {descricao_curta}')

        remover = st.selectbox("Selecione o produto para remover", options, key="remover_produto")

        if st.button("Remover produto", key="remover_botao"):
            # Encontrar o índice correto
            idx = options.index(remover)
            
            # Remover do session state
            produto_removido = st.session_state.produtos.pop(idx)
            
            # Recriar a planilha do zero
            criar_planilha_from_scratch()

            # Exibir mensagem de sucesso
            mensagem_remocao = f"🗑️ Produto '{produto_removido['Descrição'][:30]}...' removido com sucesso."
            if len(st.session_state.produtos) < LIMITE_PRODUTOS:
                mensagem_remocao += f" Agora há {LIMITE_PRODUTOS - len(st.session_state.produtos)} espaços disponíveis."
            
            st.success(mensagem_remocao)
            
            # Forçar rerun para atualizar a lista
            st.rerun()

        st.markdown("---")
        
        if st.session_state.excel_buffer:
            # Informar quantos produtos serão incluídos
            produtos_incluidos = min(len(st.session_state.produtos), LIMITE_PRODUTOS)
            st.info(f"📄 **Na planilha serão incluídos:** {produtos_incluidos} produtos")
            
            if len(st.session_state.produtos) > LIMITE_PRODUTOS:
                st.warning(f"⚠️ **Atenção:** Apenas os primeiros {LIMITE_PRODUTOS} produtos serão incluídos na planilha.")
            
            st.download_button(
                "📥 Baixar formulário Excel",
                data=st.session_state.excel_buffer,
                file_name="solicitar placa.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_final"
            )
    else:
        st.info("Nenhum produto solicitado ainda.")